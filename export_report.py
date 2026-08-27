#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把 runner.py 的运行结果（output/raw/*.json + output/summary.json）回填到报告模板，
生成一份可直接交付的对比报告 Excel。

用法：
    python export_report.py
    python export_report.py --out output --template report_template.xlsx --result report.xlsx
"""

import argparse
import datetime
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))


def percentile(values, q):
    if not values:
        return None
    s = sorted(values)
    k = (len(s) - 1) * q
    f = int(k)
    if f + 1 >= len(s):
        return round(s[f], 2)
    return round(s[f] + (s[f + 1] - s[f]) * (k - f), 2)


def load_raw_rows(raw_dir):
    rows = []
    for layer in ("performance", "compatibility", "quality", "long_context"):
        d = os.path.join(raw_dir, layer)
        if not os.path.isdir(d):
            continue
        for fn in sorted(os.listdir(d)):
            if fn.endswith(".json"):
                with open(os.path.join(d, fn), encoding="utf-8") as f:
                    r = json.load(f)
                r["_layer"] = layer
                rows.append(r)
    return rows


def build_detail_rows(rows):
    perf, compat, quality, lc = [], [], [], []
    perf_ttft = {}
    for r in rows:
        layer = r.get("_layer")
        meta = r.get("meta", {})
        m = r.get("metrics") or {}
        pid = meta.get("provider_id", "")
        if layer == "performance":
            if r.get("skipped"):
                perf.append([meta.get("case_id"), meta.get("category"), meta.get("provider_name"),
                             "", "", "", "", "", "", "", "", "", "", "", "skipped", r.get("skip_reason", "")])
            elif r.get("error"):
                perf.append([meta.get("case_id"), meta.get("category"), meta.get("provider_name"),
                             "", "", "", "", "", "", "", "", "", "", "", r["error"].get("type"), r["error"].get("message", "")])
            else:
                cint = m.get("chunk_interval_ms") or {}
                perf.append([meta.get("case_id"), meta.get("category"), meta.get("provider_name"),
                             m.get("ttft_ms"), m.get("content_ttft_ms"), m.get("e2e_ms"), m.get("generation_ms"),
                             m.get("completion_tokens"), m.get("tokens_per_sec"), m.get("reasoning_tokens"),
                             m.get("cached_tokens"), m.get("chunk_count"), cint.get("avg"), cint.get("std"), "", ""])
                if m.get("ttft_ms") is not None:
                    perf_ttft.setdefault(pid, []).append(m["ttft_ms"])
        elif layer == "compatibility":
            v = r.get("validation") or {}
            checks = v.get("checks") or []
            if r.get("error"):
                compat.append([meta.get("case_id"), meta.get("compat_feature"), meta.get("provider_name"),
                               "失败", "请求失败: %s" % r["error"].get("message", "")[:150], "", "请求失败"])
            else:
                detail = "；".join("%s：%s" % (c.get("name"), c.get("detail")) for c in checks if c.get("severity") in ("fail", "warn")) or "正常"
                compat.append([meta.get("case_id"), meta.get("compat_feature"), meta.get("provider_name"),
                               "通过" if v.get("passed") else "失败", detail, m.get("completion_tokens"), ""])
        elif layer == "quality":
            v = r.get("validation") or {}
            checks = v.get("checks") or []
            out = (r.get("output") or {}).get("text") or ""
            reference = json.dumps(meta.get("reference"), ensure_ascii=False) if meta.get("reference") else ""
            detail = "；".join("%s：%s" % (c.get("name"), c.get("detail")) for c in checks)
            quality.append([meta.get("case_id"), meta.get("category"), meta.get("provider_name"),
                            "通过" if v.get("passed") else "失败", detail, reference, out[:120], ""])
        elif layer == "long_context":
            v = r.get("validation") or {}
            checks = v.get("checks") or []
            out = (r.get("output") or {}).get("text") or ""
            if r.get("skipped"):
                lc.append([meta.get("case_id"), meta.get("provider_name"), meta.get("target_tokens"), "跳过",
                           meta.get("depth"), "", "", r.get("skip_reason", "")])
            elif r.get("error"):
                lc.append([meta.get("case_id"), meta.get("provider_name"), meta.get("target_tokens"), "错误",
                           meta.get("depth"), "", "", r["error"].get("message", "")[:150]])
            else:
                detail = "；".join("%s：%s" % (c.get("name"), c.get("detail")) for c in checks)
                lc.append([meta.get("case_id"), meta.get("provider_name"), meta.get("target_tokens"),
                           "是" if v.get("passed") else "否", meta.get("depth"), out[:80], m.get("completion_tokens"), detail])
    return perf, compat, quality, lc, perf_ttft


def fill_sheet(ws, rows, num_fmts):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, rv in enumerate(rows, 2):
        for c, v in enumerate(rv, 1):
            if v == "":
                v = None
            cell = ws.cell(row=i, column=c, value=v)
            if num_fmts and c <= len(num_fmts) and num_fmts[c - 1]:
                cell.number_format = num_fmts[c - 1]


def resolve_run_dir(out):
    """若 out 直接含 summary.json 则原样返回；否则定位其下最新的完整运行目录（含 summary.json）。"""
    if os.path.exists(os.path.join(out, "summary.json")):
        return out
    subs = sorted(d for d in os.listdir(out)
                  if d.startswith("run_") and os.path.isdir(os.path.join(out, d)))
    for d in reversed(subs):  # 跳过仅 --sample 的目录
        if os.path.exists(os.path.join(out, d, "summary.json")):
            return os.path.join(out, d)
    return os.path.join(out, subs[-1]) if subs else out


def rebuild_summary_from_raw(raw_dir):
    """没有 summary.json 时，从 raw/*.json 反推一份最小 summary（提供 providers/环境/时间）。"""
    rows = load_raw_rows(raw_dir)
    providers = {}
    env = None
    min_sent = max_sent = None
    for r in rows:
        meta = r.get("meta") or {}
        pid = meta.get("provider_id")
        if pid and pid not in providers:
            providers[pid] = {"id": pid, "name": meta.get("provider_name") or pid,
                              "model": meta.get("model", ""), "base_url": meta.get("base_url", "")}
        if env is None and r.get("environment"):
            env = r["environment"]
        sent = meta.get("sent_at")
        if sent:
            min_sent = sent if min_sent is None else min(min_sent, sent)
            max_sent = sent if max_sent is None else max(max_sent, sent)
    return {
        "providers": list(providers.values()),
        "environment": env or {},
        "started_at": min_sent or "",
        "finished_at": max_sent or "",
        "test_config": {},
        "rebuilt_from_raw": True,
    }


def generate_report(out_dir, template_path, result_path):
    """把一次完整运行的原始数据回填到报告模板，生成 Excel。返回生成的文件路径。"""
    summary_path = os.path.join(out_dir, "summary.json")
    if os.path.exists(summary_path):
        with open(summary_path, encoding="utf-8") as f:
            summary = json.load(f)
    else:
        summary = rebuild_summary_from_raw(os.path.join(out_dir, "raw"))
        print("[提示] 未找到 summary.json，已从 raw 兜底重建（参数配置/评分信息可能不完整）")

    rows = load_raw_rows(os.path.join(out_dir, "raw"))
    perf, compat, quality, lc, perf_ttft = build_detail_rows(rows)

    wb = openpyxl.load_workbook(template_path)
    fill_sheet(wb["性能明细"], perf, [None, None, None, "0.0", "0.0", "0.0", "0.0", "0", "0.00", "0", "0", "0", "0.0", "0.0", None, None])
    fill_sheet(wb["兼容性明细"], compat, [None, None, None, None, None, "0", None])
    fill_sheet(wb["质量明细"], quality, [None, None, None, None, None, None, None, None])
    fill_sheet(wb["长上下文明细"], lc, [None, None, "0", None, "0.00", None, "0", None])

    ws = wb["汇总表"]
    providers = summary.get("providers", [])
    for i in range(2, 22):
        ws.cell(row=i, column=1).value = None
        ws.cell(row=i, column=3).value = None
        ws.cell(row=i, column=4).value = None
    for i, p in enumerate(providers):
        row = i + 2
        ws.cell(row=row, column=1, value=p.get("name"))
        ttfts = perf_ttft.get(p.get("id"), [])
        ws.cell(row=row, column=3, value=percentile(ttfts, 0.5))
        ws.cell(row=row, column=4, value=percentile(ttfts, 0.95))

    ws2 = wb["评分模型"]
    for i in range(2, 22):
        ws2.cell(row=i, column=1).value = None
    for i, p in enumerate(providers):
        ws2.cell(row=i + 2, column=1, value=p.get("name"))

    ws0 = wb["说明"]
    ws0["A2"] = "本次运行信息"
    env = summary.get("environment") or {}
    info = (
        "运行时间：%s ~ %s\n"
        "环境标注：%s\n"
        "主机：%s | OS：%s | 架构：%s\n"
        "CPU：%s 核 | 内存：%s GB | Python：%s | 时区：%s\n"
        "渠道数：%d | 输出目录：%s\n"
        "%s"
        "数据来源：runner.py 原始结果，已回填至本报告。"
    ) % (
        summary.get("started_at", ""), summary.get("finished_at", ""),
        env.get("env_tag", ""), env.get("hostname", ""), env.get("os", ""), env.get("machine", ""),
        env.get("cpu_cores", ""), env.get("memory_gb", ""), env.get("python", ""), env.get("tz", ""),
        len(providers), os.path.abspath(out_dir),
        "⚠️ 本报告由 raw 兜底重建（原 summary.json 缺失），评分权重/参数配置信息不完整。\n" if summary.get("rebuilt_from_raw") else "")
    ws0.cell(row=3, column=2, value=info).alignment = Alignment(vertical="top", wrap_text=True)

    fill_experiment_config(wb, summary)

    # 合并缓存测试结果（若本次运行了 --cache）
    cache_path = os.path.join(out_dir, "cache_summary.json")
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            cache_s = json.load(f)
        ws_cache = wb.create_sheet("缓存命中")
        fill_cache_sheet(ws_cache, cache_s)

    wb.save(result_path)
    return result_path


def generate_sample_report(out_dir):
    """为 --sample 运行生成一张简单对比 Excel。"""
    sample_path = os.path.join(out_dir, "sample_summary.json")
    if not os.path.exists(sample_path):
        return None
    with open(sample_path, encoding="utf-8") as f:
        s = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "示例结果"
    headers = ["模型", "渠道", "TTFT(ms)", "E2E(ms)", "速度(tok/s)", "内容字数", "推理字数", "状态", "输出预览"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([18, 12, 10, 10, 12, 9, 9, 10, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 2
    for row in s.get("rows", []):
        m = row.get("metrics") or {}
        status = "错误" if row.get("error") else ("跳过" if row.get("skipped") else "成功")
        for c, v in enumerate([row.get("model_id"), row.get("channel_name"), m.get("ttft_ms"),
                               m.get("e2e_ms"), m.get("tokens_per_sec"), m.get("content_chars"),
                               m.get("reasoning_chars"), status, row.get("output_preview", "")], 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    env = s.get("environment") or {}
    ws.cell(row=r + 1, column=1, value="环境标注：%s | 主机：%s | OS：%s" % (
        env.get("env_tag", ""), env.get("hostname", ""), env.get("os", "")))
    result_path = os.path.join(out_dir, "sample_report.xlsx")
    wb.save(result_path)
    return result_path


CACHE_METHOD = (
    "缓存测试方法：长前缀=确定性生成2000 token文本(各渠道一致)；"
    "①连打3次=同前缀+同问题连发3次(间隔1s)，看第2/3次 cached_tokens>0 且 TTFT 下降；"
    "②前缀复用=同前缀+不同问题，看 cached_tokens≈前缀token数；"
    "③基线=完全不同前缀，应 cached_tokens=0(防误判)。判据：cached_tokens>0=命中；命中率=第2/3次命中次数/2。"
)


def fill_cache_sheet(ws, s):
    """把缓存测试结果写进一个 worksheet（含测试方法说明）。"""
    headers = ["模型", "渠道", "第1次TTFT(ms)", "第2次TTFT(ms)", "第3次TTFT(ms)",
               "cached(第2/3次)", "前缀cached", "基线cached", "命中率", "是否命中"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([16, 12, 12, 12, 12, 14, 12, 12, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    def ttft(sp, i):
        return sp[i - 1].get("ttft_ms") if len(sp) >= i else None

    def cached(sp, i):
        return sp[i - 1].get("cached_tokens") if len(sp) >= i else None

    r = 2
    for row in s.get("rows", []):
        sp = row.get("same_prompt", [])
        pr = row.get("prefix_reuse", {}).get("cached_tokens")
        bl = row.get("baseline", {}).get("cached_tokens")
        vals = [row.get("model_id"), row.get("channel_name"), ttft(sp, 1), ttft(sp, 2), ttft(sp, 3),
                "%s / %s" % (cached(sp, 2), cached(sp, 3)), pr, bl,
                row.get("hit_rate", ""), "是" if row.get("hit") else "否"]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    env = s.get("environment") or {}
    ws.cell(row=r + 1, column=1, value="前缀 %d token | 连打 %d 次 | 环境标注：%s | 主机：%s" % (
        s.get("prefix_tokens", 0), s.get("runs", 0), env.get("env_tag", ""), env.get("hostname", "")))
    mcell = ws.cell(row=r + 3, column=1, value=CACHE_METHOD)
    mcell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 4, end_column=10)


def fill_experiment_config(wb, summary):
    """把测试参数/渠道/模型/用例设计写成一个「实验配置」sheet。"""
    tc = summary.get("test_config") or {}
    ws = wb["实验配置"] if "实验配置" in wb.sheetnames else wb.create_sheet("实验配置")
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 78
    ws["A1"] = "实验配置（测试参数 + 渠道/模型 + 用例设计）"
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    if not tc:
        ws.cell(row=r, column=1, value="（本报告由 raw 兜底重建，无完整测试参数配置；可重新完整运行 runner.py 以补齐此页）")
        r += 2

    def sec(title, pairs):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
        r += 1
        for k, v in pairs:
            ws.cell(row=r, column=1, value=k)
            c = ws.cell(row=r, column=2, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1

    sp = tc.get("shared_request_params") or {}
    to = tc.get("timeouts") or {}
    w = tc.get("score_weights") or {}
    sec("1. 共享请求参数", [
        ("temperature", sp.get("temperature", "不传")),
        ("top_p", sp.get("top_p", tc.get("top_p_note", "默认不传"))),
    ])
    sec("2. 推理 / 超时 / 重试", [
        ("reasoning_max_tokens（推理模型下限）", tc.get("reasoning_max_tokens")),
        ("超时（连接/读/长上下文）", "%ss / %ss / %ss" % (
            to.get("connect_seconds", 10), to.get("read_seconds", 180), to.get("read_seconds_long_context", 600))),
        ("失败重试次数", tc.get("retries")),
    ])
    sec("3. 评分权重", [(k, v) for k, v in w.items()])
    sec("4. 渠道", [(ch.get("name", ch.get("id")), ch.get("base_url")) for ch in tc.get("channels", [])])
    sec("5. 模型", [(m.get("id"), "aliases=%s | 推理模型=%s | 标称上下文=%s | supports=%s" % (
        json.dumps(m.get("aliases", {}), ensure_ascii=False),
        "是" if m.get("reasoning") else "否", m.get("context"),
        json.dumps(m.get("supports", {}), ensure_ascii=False))) for m in tc.get("models", [])])
    sec("6. 用例设计", [(layer, "%d 条 | %s" % (info.get("count", 0), info.get("desc", "")))
                        for layer, info in tc.get("layers", {}).items()])
    sec("7. 运行环境", [(k, v) for k, v in (summary.get("environment") or {}).items()])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(HERE, "output"))
    ap.add_argument("--template", default=os.path.join(HERE, "report_template.xlsx"))
    ap.add_argument("--result", default="")
    args = ap.parse_args()

    out_dir = resolve_run_dir(args.out)
    result_path = args.result or os.path.join(HERE, "report_%s.xlsx" % datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    generate_report(out_dir, args.template, result_path)
    print("已生成报告：%s" % os.path.abspath(result_path))


if __name__ == "__main__":
    main()
