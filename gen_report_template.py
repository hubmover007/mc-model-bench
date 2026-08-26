#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成多渠道测试报告模板 Excel（含指标对比表 + 评分公式）。

Sheet：说明 / 汇总表 / 性能明细 / 兼容性明细 / 质量明细 / 长上下文明细 / 评分模型
用法：python gen_report_template.py  ->  report_template.xlsx
"""

import json
import os
import random

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "report_template.xlsx")
MAX_PROVIDERS = 20

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, num_fmts=None):
    for c, v in enumerate(values, 1):
        if v == "":
            v = None
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = BORDER
        if num_fmts and c <= len(num_fmts) and num_fmts[c - 1]:
            cell.number_format = num_fmts[c - 1]
    return row + 1


def load_cases(layer):
    path = os.path.join(HERE, "test_cases", "%s.json" % layer)
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return []


def build_example():
    rng = random.Random(42)
    providers = ["EasyRouter(示例)", "Mayi(示例)"]
    perf_cases = load_cases("performance")
    compat_cases = load_cases("compatibility")
    quality_cases = load_cases("quality")
    lc_cases = load_cases("long_context")

    profile = {
        "短输入短输出": ((250, 600), (500, 1500), (40, 90)),
        "短输入长输出": ((350, 800), (8000, 25000), (900, 1600)),
        "长输入短输出": ((1200, 3500), (3000, 8000), (60, 90)),
        "长输入长输出": ((1500, 4000), (12000, 30000), (1000, 1400)),
    }
    perf_rows = []
    for pidx, prov in enumerate(providers):
        for c in perf_cases:
            cat = c.get("category") or "短输入短输出"
            lo_t, hi_t = profile.get(cat, profile["短输入短输出"])[0]
            lo_e, hi_e = profile.get(cat, profile["短输入短输出"])[1]
            lo_k, hi_k = profile.get(cat, profile["短输入短输出"])[2]
            slow = 1.0 + 0.35 * pidx
            ttft = rng.randint(int(lo_t * slow), int(hi_t * slow))
            e2e = rng.randint(int(lo_e * slow), int(hi_e * slow))
            tokens = rng.randint(lo_k, hi_k)
            gen = max(50, e2e - ttft)
            speed = round(tokens / (gen / 1000.0), 2)
            reasoning = rng.randint(200, 900)
            cached = rng.choice([0, 0, 120])
            chunk = rng.randint(8, 40)
            interval = round(gen / chunk, 1)
            std = round(interval * rng.uniform(0.3, 0.8), 1)
            perf_rows.append([c["id"], cat, prov, ttft, round(ttft * 1.2), e2e, gen, tokens, speed,
                              reasoning, cached, chunk, interval, std, "", ""])
    perf_rows.append(["perf_lis_03", "长输入短输出", "Mayi(示例)", "", "", "", "", "", "", "", "",
                      "", "", "", "timeout", "读超时（>180s）"])

    compat_rows = []
    fail_map = {"Mayi(示例)": {"cmp_tools_01": "未返回 tool_calls，模型直接以文本回答", "cmp_json_01": "输出非 JSON，无法 json.loads"}}
    for prov in providers:
        for c in compat_cases:
            fail = fail_map.get(prov, {}).get(c["id"])
            compat_rows.append([c["id"], c.get("compat_feature"), prov, "失败" if fail else "通过",
                                fail or "", rng.randint(20, 300), ""])

    quality_rows = []
    for prov in providers:
        for i, c in enumerate(quality_cases):
            passed = (prov == "EasyRouter(示例)") or (i % 4 != 0)
            quality_rows.append([c["id"], c.get("category"), prov, "通过" if passed else "失败",
                                 "judge 命中" if passed else "judge 未命中", json.dumps(c.get("reference"), ensure_ascii=False),
                                 "", ""])

    lc_rows = []
    for c in lc_cases:
        lc_rows.append([c["id"], "EasyRouter(示例)", c.get("target_tokens"), "是", (c.get("needle") or {}).get("depth"),
                        "抽奖码（检索成功）", 24, ""])
    for c in lc_cases:
        if c.get("target_tokens") == 128000:
            result, note = "跳过", "超出标称上下文（64K）"
        elif c.get("target_tokens") == 64000:
            result, note = "否", "未检索到"
        else:
            result, note = "是", ""
        lc_rows.append([c["id"], "Mayi(示例)", c.get("target_tokens"), result, (c.get("needle") or {}).get("depth"),
                        "抽奖码（检索成功）" if result == "是" else "（未命中）", 24, note])

    return providers, perf_rows, compat_rows, quality_rows, lc_rows


def build_workbook():
    providers, perf_rows, compat_rows, quality_rows, lc_rows = build_example()
    wb = openpyxl.Workbook()

    # ---------- 说明 ----------
    ws = wb.active
    ws.title = "说明"
    set_widths(ws, [16, 110])
    ws["A1"] = "多渠道模型测试报告（模板）"
    ws["A1"].font = TITLE_FONT
    notes = [
        ("用途", "统一对比“同一模型 × 不同渠道”的服务能力。汇总表与评分模型的公式自动从四张明细表取值。"),
        ("1. 明细表", "性能明细：TTFT/速度/E2E/推理tokens/缓存tokens/chunk分布；错误行只填“错误类型/错误详情”，其余留空（公式自动排除）。\n"
                       "兼容性明细：结果填 通过/失败，失败表现记录具体现象（未返回 tool_calls、输出非 JSON 等）。\n"
                       "质量明细：来自 HuggingFace（GSM8K/TruthfulQA）的判分结果。\n"
                       "长上下文明细：检索成功填 是/否/跳过。"),
        ("2. 汇总表", "TTFT均值/速度/E2E/错误率/兼容率/质量正确率/检索率均为公式；TTFT P50/P95 由 export_report.py 计算写入。"),
        ("3. 评分模型", "性能得分=平均(TTFT,速度,E2E 得分)；兼容得分=通过率×100；质量得分=正确率×100；长上下文得分=检索率×100；\n"
                       "总评分=性能×0.3+兼容×0.3+质量×0.2+长上下文×0.2（权重在评分模型表 L2:M5 可改）。"),
        ("注意", "模板预填了两条示例渠道数据用于演示公式，正式使用请用 export_report.py 回填真实结果。"),
    ]
    r = 2
    for k, v in notes:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True) if k else Font()
        ws.cell(row=r, column=2, value=v).alignment = LEFT
        r += 1

    # ---------- 汇总表 ----------
    ws = wb.create_sheet("汇总表")
    headers = ["渠道", "TTFT均值(ms)", "TTFT P50(ms)", "TTFT P95(ms)", "速度均值(tok/s)", "E2E均值(ms)",
               "错误率(%)", "兼容通过率(%)", "质量正确率(%)", "长上下文检索率(%)", "备注"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [16, 12, 12, 12, 12, 12, 10, 12, 12, 13, 28])
    ws.freeze_panes = "A2"
    num_fmts = [None, "0.0", "0.0", "0.0", "0.00", "0.0", "0.0", "0.0", "0.0", "0.0", None]
    for i in range(MAX_PROVIDERS):
        row = i + 2
        a = "$A%d" % row
        ws.cell(row=row, column=1, value=providers[i] if i < len(providers) else None)
        formulas = [
            '=IFERROR(ROUND(AVERAGEIFS(性能明细!$D$2:$D$2000,性能明细!$C$2:$C$2000,%s,性能明细!$O$2:$O$2000,""),1),"")' % a,
            None, None,
            '=IFERROR(ROUND(AVERAGEIFS(性能明细!$I$2:$I$2000,性能明细!$C$2:$C$2000,%s,性能明细!$O$2:$O$2000,""),2),"")' % a,
            '=IFERROR(ROUND(AVERAGEIFS(性能明细!$F$2:$F$2000,性能明细!$C$2:$C$2000,%s,性能明细!$O$2:$O$2000,""),1),"")' % a,
            '=IFERROR(ROUND(COUNTIFS(性能明细!$C$2:$C$2000,%s,性能明细!$O$2:$O$2000,"error")/COUNTIF(性能明细!$C$2:$C$2000,%s)*100,1),"")' % (a, a),
            '=IFERROR(ROUND(COUNTIFS(兼容性明细!$C$2:$C$2000,%s,兼容性明细!$D$2:$D$2000,"通过")/COUNTIF(兼容性明细!$C$2:$C$2000,%s)*100,1),"")' % (a, a),
            '=IFERROR(ROUND(COUNTIFS(质量明细!$C$2:$C$2000,%s,质量明细!$D$2:$D$2000,"通过")/COUNTIF(质量明细!$C$2:$C$2000,%s)*100,1),"")' % (a, a),
            '=IFERROR(ROUND(COUNTIFS(长上下文明细!$B$2:$B$2000,%s,长上下文明细!$D$2:$D$2000,"是")/COUNTIFS(长上下文明细!$B$2:$B$2000,%s,长上下文明细!$D$2:$D$2000,"<>跳过")*100,1),"")' % (a, a),
        ]
        for col in range(2, len(headers)):
            f = formulas[col - 2]
            if f is not None:
                ws.cell(row=row, column=col, value=f)
            if num_fmts[col - 1]:
                ws.cell(row=row, column=col).number_format = num_fmts[col - 1]
            ws.cell(row=row, column=col).border = BORDER
    ws.conditional_formatting.add("B2:B%d" % (MAX_PROVIDERS + 1),
                                  ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                                                 mid_type="num", mid_value=1500, mid_color="FFEB84",
                                                 end_type="num", end_value=4000, end_color="F8696B"))

    # ---------- 性能明细 ----------
    ws = wb.create_sheet("性能明细")
    headers = ["用例ID", "分类", "渠道", "TTFT(ms)", "内容TTFT(ms)", "E2E(ms)", "生成耗时(ms)", "输出tokens",
               "速度(tok/s)", "推理tokens", "缓存tokens", "chunk数", "chunk间隔均值(ms)", "chunk间隔std", "错误类型", "错误详情"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [14, 14, 15, 10, 11, 10, 11, 10, 11, 11, 11, 9, 14, 12, 12, 40])
    ws.freeze_panes = "A2"
    num_fmts = [None, None, None, "0.0", "0.0", "0.0", "0.0", "0", "0.00", "0", "0", "0", "0.0", "0.0", None, None]
    for i, rv in enumerate(perf_rows, 2):
        write_row(ws, i, rv, num_fmts)

    # ---------- 兼容性明细 ----------
    ws = wb.create_sheet("兼容性明细")
    headers = ["用例ID", "功能点", "渠道", "结果", "失败表现/详情", "输出tokens", "备注"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [16, 18, 15, 10, 52, 11, 22])
    ws.freeze_panes = "A2"
    num_fmts = [None, None, None, None, None, "0", None]
    for i, rv in enumerate(compat_rows, 2):
        write_row(ws, i, rv, num_fmts)

    # ---------- 质量明细 ----------
    ws = wb.create_sheet("质量明细")
    headers = ["用例ID", "分类", "渠道", "结果", "判分详情", "参考答案", "输出摘要", "备注"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [18, 12, 15, 10, 26, 40, 40, 22])
    ws.freeze_panes = "A2"
    num_fmts = [None, None, None, None, None, None, None, None]
    for i, rv in enumerate(quality_rows, 2):
        write_row(ws, i, rv, num_fmts)

    # ---------- 长上下文明细 ----------
    ws = wb.create_sheet("长上下文明细")
    headers = ["用例ID", "渠道", "梯度(tokens)", "检索成功", "针位置", "答案摘要", "输出tokens", "备注"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [16, 15, 12, 10, 9, 30, 11, 28])
    ws.freeze_panes = "A2"
    num_fmts = [None, None, "0", None, "0.00", None, "0", None]
    for i, rv in enumerate(lc_rows, 2):
        write_row(ws, i, rv, num_fmts)

    # ---------- 评分模型 ----------
    ws = wb.create_sheet("评分模型")
    headers = ["渠道", "TTFT得分", "速度得分", "E2E得分", "性能得分", "兼容得分", "质量得分", "长上下文得分", "总评分", "排名"]
    write_row(ws, 1, headers)
    style_header(ws, len(headers))
    set_widths(ws, [16, 10, 10, 10, 10, 10, 10, 12, 10, 8])
    ws.freeze_panes = "A2"
    ws["L1"] = "权重区（可修改）↓"
    ws["L1"].font = Font(bold=True, color="C00000")
    weights = [("性能权重", 0.3), ("兼容权重", 0.3), ("质量权重", 0.2), ("长上下文权重", 0.2)]
    for j, (lab, val) in enumerate(weights, start=2):
        ws.cell(row=j, column=12, value=lab).font = Font(bold=True)
        ws.cell(row=j, column=12).alignment = CENTER
        ws.cell(row=j, column=13, value=val).number_format = "0.0"
        ws.cell(row=j, column=13).border = BORDER
    num_fmts = [None, "0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "0"]
    for i in range(MAX_PROVIDERS):
        row = i + 2
        ws.cell(row=row, column=1, value=providers[i] if i < len(providers) else None)
        formulas = {
            2: '=IFERROR(ROUND(100*MIN(汇总表!$B$2:$B$%d)/汇总表!B%d,1),"")' % (MAX_PROVIDERS + 1, row),
            3: '=IFERROR(ROUND(100*汇总表!E%d/MAX(汇总表!$E$2:$E$%d),1),"")' % (row, MAX_PROVIDERS + 1),
            4: '=IFERROR(ROUND(100*MIN(汇总表!$F$2:$F$%d)/汇总表!F%d,1),"")' % (MAX_PROVIDERS + 1, row),
            5: '=IFERROR(ROUND(AVERAGE(B%d:D%d),1),"")' % (row, row),
            6: '=IFERROR(ROUND(汇总表!H%d,1),"")' % row,
            7: '=IFERROR(ROUND(汇总表!I%d,1),"")' % row,
            8: '=IFERROR(ROUND(汇总表!J%d,1),"")' % row,
            9: '=IF(OR(ISNUMBER(E%d),ISNUMBER(F%d),ISNUMBER(G%d),ISNUMBER(H%d)),ROUND(IFERROR(E%d,0)*$M$2+IFERROR(F%d,0)*$M$3+IFERROR(G%d,0)*$M$4+IFERROR(H%d,0)*$M$5,1),"")' % (row, row, row, row, row, row, row, row),
            10: '=IF(ISNUMBER(I%d),RANK(I%d,$I$2:$I$%d),"")' % (row, row, MAX_PROVIDERS + 1),
        }
        for col in range(2, len(headers) + 1):
            f = formulas.get(col)
            cell = ws.cell(row=row, column=col)
            if f is not None:
                cell.value = f
            if num_fmts[col - 1]:
                cell.number_format = num_fmts[col - 1]
            cell.border = BORDER
    ws.conditional_formatting.add("I2:I%d" % (MAX_PROVIDERS + 1),
                                  ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                                                 mid_type="num", mid_value=50, mid_color="FFEB84",
                                                 end_type="num", end_value=100, end_color="63BE7B"))

    wb.save(OUT_PATH)
    print("已生成：%s" % os.path.abspath(OUT_PATH))


if __name__ == "__main__":
    build_workbook()
